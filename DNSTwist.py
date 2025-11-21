import os, csv, json, re, time, requests, subprocess
from io import BytesIO
from collections import deque

import dns.resolver
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from PIL import Image

# ---------------- CONFIG ----------------
# Define your input andoutput file names 👇
INPUT_CSV   = "twist.csv"
OUTPUT_XLSX = "dnstwist_results.xlsx"

# Put your API key here 👇
URLSCAN_API_KEY    = "<URL SCAN API KEY HERE>"
URLSCAN_VISIBILITY = "private"   # can be "private", "unlisted", or "public"

# API / request safety limits
MAX_PER_MINUTE = 120
_req_times_minute = deque()

# -----------------------------------------

def rate_limit_per_minute():
    """Soft limiter for URLScan API"""
    now = time.monotonic()
    while _req_times_minute and now - _req_times_minute[0] > 60.0:
        _req_times_minute.popleft()
    if len(_req_times_minute) >= MAX_PER_MINUTE:
        sleep_for = max(0.05, 60.0 - (now - _req_times_minute[0]) + 0.05)
        print(f"[Rate-limit] Sleeping {sleep_for:.1f}s to respect API limit...")
        time.sleep(sleep_for)
    _req_times_minute.append(time.monotonic())

# -----------------------------------------
# CSV loader
# -----------------------------------------
def load_domains_from_csv(path: str):
    domains = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        pos = f.tell()
        reader = csv.DictReader(f)
        headers = [h.strip().lower() for h in (reader.fieldnames or [])] if reader.fieldnames else []
        if "domain" in headers:
            for row in reader:
                d = (row.get("domain") or "").strip()
                if d:
                    domains.append(d)
        else:
            f.seek(pos)
            r = csv.reader(f)
            for row in r:
                if not row: continue
                d = (row[0] or "").strip()
                if d and d.lower() != "domain": domains.append(d)
    seen, out = set(), []
    for d in domains:
        if d not in seen:
            seen.add(d)
            out.append(d)
    return out

# -----------------------------------------
# dnstwist runner
# -----------------------------------------
def run_dnstwist(domain: str, use_registered: bool = True):
    base = ["dnstwist", "--whois", "--geoip", "--mxcheck", "--format", "json"]
    if use_registered: base.insert(1, "--registered")
    cmd = base + [domain]
    try:
        proc = subprocess.run(cmd, capture_output=True, text=True, timeout=600)
    except Exception as e:
        print(f"[ERROR] launching dnstwist: {e}")
        return []
    if proc.returncode != 0 and not proc.stdout:
        print(f"[WARN] dnstwist exit={proc.returncode}, stderr={proc.stderr.strip()}")
        return []
    text = proc.stdout.strip()
    m = re.search(r"\[\s*{.*}\s*\]\s*$", text, flags=re.S)
    json_text = m.group(0) if m else text
    try:
        data = json.loads(json_text)
        if isinstance(data, dict): data = [data]
        return data if isinstance(data, list) else []
    except json.JSONDecodeError as je:
        print(f"[WARN] JSON decode error: {je}; head: {text[:300]}")
        return []

# -----------------------------------------
# DNS and WHOIS enrichment
# -----------------------------------------
_resolver = dns.resolver.Resolver(configure=True)
_resolver.nameservers = ["1.1.1.1", "8.8.8.8"]
_resolver.lifetime = 3.0
_resolver.timeout  = 2.0

def resolve_records(domain: str):
    out = {"A": [], "AAAA": [], "NS": [], "MX": []}
    def q(name, rtype):
        try:
            ans = _resolver.resolve(name, rtype, raise_on_no_answer=False)
            return [r.to_text() for r in ans] if ans else []
        except Exception:
            return []
    out["A"]    = q(domain, "A")
    out["AAAA"] = q(domain, "AAAA")
    out["NS"]   = q(domain, "NS")
    out["MX"]   = q(domain, "MX")
    return out

def whois_full(domain: str, timeout=25):
    try:
        proc = subprocess.run(["whois", domain], capture_output=True, text=True, timeout=timeout)
        txt = proc.stdout.strip()
        txt = re.sub(r"\n{3,}", "\n\n", txt)
        return txt
    except Exception as e:
        return f"[whois error] {e}"

# -----------------------------------------
# URLScan.io interaction
# -----------------------------------------
def urlscan_submit(url: str):
    if not URLSCAN_API_KEY: return None
    rate_limit_per_minute()
    r = requests.post(
        "https://urlscan.io/api/v1/scan/",
        headers={"API-Key": URLSCAN_API_KEY, "Content-Type": "application/json"},
        json={"url": url, "visibility": URLSCAN_VISIBILITY},
        timeout=30,
    )
    if r.status_code == 200:
        return r.json().get("uuid")
    if r.status_code == 400:
        return None
    r.raise_for_status()
    return None

def urlscan_poll(uuid: str, retries=18, delay=5):
    headers = {"API-Key": URLSCAN_API_KEY}
    url = f"https://urlscan.io/api/v1/result/{uuid}/"
    for _ in range(retries):
        rate_limit_per_minute()
        r = requests.get(url, headers=headers, timeout=30)
        if r.status_code == 200:
            return r.json()
        if r.status_code == 404:
            time.sleep(delay)
            continue
        r.raise_for_status()
    return None

def fetch_screenshot_bytes(result_json: dict):
    if not result_json: return "", None
    task = result_json.get("task") or {}
    screenshot_url = task.get("screenshotURL", "")
    report_url     = task.get("reportURL", "")
    if not screenshot_url:
        return report_url, None
    rate_limit_per_minute()
    r = requests.get(screenshot_url, timeout=30)
    if r.status_code == 200:
        return report_url, r.content
    return report_url, None

# -----------------------------------------
# Excel handling
# -----------------------------------------
HEADERS = [
    "input_domain","permutation","fuzzer",
    "dns_a","dns_aaaa","dns_ns","mx","geoip",
    "whois_created","whois_updated","whois_registrar",
    "urlscan_report","screenshot","whois_full"
]

def init_wb():
    wb = Workbook()
    ws = wb.active
    ws.title = "results"
    ws.append(HEADERS)
    widths = [20,30,12,28,28,28,28,18,18,18,26,36,20,80]
    for i in range(1, len(HEADERS)+1):
        c = ws.cell(row=1, column=i)
        c.font = Font(bold=True)
        c.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = widths[i-1]
    return wb, ws

def list_str(v):
    return ", ".join(v) if isinstance(v, list) else (v or "")

def add_row(ws, row_idx, rd, png_bytes):
    ws.append([
        rd["input_domain"], rd["permutation"], rd["fuzzer"],
        rd["dns_a"], rd["dns_aaaa"], rd["dns_ns"], rd["mx"], rd["geoip"],
        rd["whois_created"], rd["whois_updated"], rd["whois_registrar"],
        rd["urlscan_report"], "", rd["whois_full"]
    ])
    if png_bytes:
        img = Image.open(BytesIO(png_bytes))
        MAX_W = 320
        if img.width > MAX_W:
            new_h = int(img.height * (MAX_W / float(img.width)))
            img = img.resize((MAX_W, new_h))
        buf = BytesIO()
        img.save(buf, format="PNG", optimize=True)
        buf.seek(0)
        xl_img = XLImage(buf)
        xl_img.anchor = f"{get_column_letter(13)}{row_idx}"
        ws.add_image(xl_img)
        ws.row_dimensions[row_idx].height = xl_img.height * 0.75
    for col in range(1, len(HEADERS)+1):
        ws.cell(row=row_idx, column=col).alignment = Alignment(wrap_text=True, vertical="top")

# -----------------------------------------
# Main logic
# -----------------------------------------
def main():
    if not os.path.exists(INPUT_CSV):
        print(f"[ERROR] {INPUT_CSV} not found")
        return

    bases = load_domains_from_csv(INPUT_CSV)
    if not bases:
        print("[ERROR] No domains loaded")
        return

    wb, ws = init_wb()
    row, total = 2, 0

    for base in bases:
        print(f"\n=== Base: {base} ===")
        permlist = run_dnstwist(base, use_registered=True)
        if not permlist:
            permlist = run_dnstwist(base, use_registered=False)
        print(f"[INFO] permutations: {len(permlist)}")

        for p in permlist:
            perm = p.get("domain") or p.get("idn") or ""
            dns = resolve_records(perm)
            whois_txt = whois_full(perm)

            report_url, png = "", None
            if URLSCAN_API_KEY:
                for scheme in ("https", "http"):
                    uuid = urlscan_submit(f"{scheme}://{perm}/")
                    if not uuid:
                        continue
                    res = urlscan_poll(uuid)
                    if res:
                        report_url, png = fetch_screenshot_bytes(res)
                        if png: break

            rd = {
                "input_domain": base,
                "permutation": perm,
                "fuzzer": p.get("fuzzer",""),
                "dns_a": list_str(dns["A"]),
                "dns_aaaa": list_str(dns["AAAA"]),
                "dns_ns": list_str(dns["NS"]),
                "mx": list_str(dns["MX"]),
                "geoip": p.get("geoip",""),
                "whois_created": p.get("whois_created",""),
                "whois_updated": p.get("whois_updated",""),
                "whois_registrar": p.get("whois_registrar",""),
                "urlscan_report": report_url,
                "whois_full": whois_txt,
            }

            add_row(ws, row, rd, png)
            row += 1
            total += 1

    wb.save(OUTPUT_XLSX)
    print(f"\n✅ Saved {OUTPUT_XLSX} with {total} rows.")

if __name__ == "__main__":
    main()

